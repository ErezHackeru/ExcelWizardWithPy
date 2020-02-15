# The best library for excel is openpyxl
import openpyxl
import time
from datetime import datetime

# open a workbook of excel:
wb = openpyxl.load_workbook('D:/Documents/hiSky/WizardExample.xlsx')
print(wb.sheetnames)

# open a worksheet to fill and read data
sheet1 = wb['Sheet1']
print(sheet1['B4'].value)
print(sheet1['A1'].value) # Get the value from the cell.

# change data inside the file:
sheet1['B4'].value = 'Clemantinot'

# formats of diffrent date and time to determine the file name:
my_date = datetime.now().date()
t = time.localtime()
current_time = time.strftime("%H_%M_%S", t)
#wb.save(f'D:/Documents/hiSky/WizardExample{my_date}_{current_time}.xlsx')

now = datetime.now().time() # time object
my_date = datetime.now().date()
print("my date = ", my_date)
print("now =", now)
print("type(now) =", type(now))

t = time.localtime()
current_time = time.strftime("%H_%M_%S", t)
print("current_time is", current_time)

print(f'{my_date}_{current_time}')

sheet1.title = 'My Sheet'
#wb.save(f'D:/Documents/hiSky/WizardExample{my_date}_{current_time}.xlsx')

# read or write to a few cells at once:
for i in range(1, 5):
    print(sheet1.cell(row = i, column = 3).value)

print(f'max row is: {sheet1.max_row}')
print(f'max column is: {sheet1.max_column}')
print(f'leeter fo column 1: {openpyxl.utils.get_column_letter(1)}')
print(f'leeter fo column 6: {openpyxl.utils.get_column_letter(6)}')
print(f'leeter fo column 26: {openpyxl.utils.get_column_letter(26)}')
print('')
for rowOfCellObjects in sheet1['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')